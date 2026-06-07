"""
Gantt Chart — isolated file parsers for draft extraction (Phase 3).
No CMP coupling; each parser returns GanttDraftTask-shaped dicts.
"""
import csv
import json
import logging
import os
import re
import uuid
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import fitz  # PyMuPDF

from services.gantt_constants import GANTT_EXTRACTION_MODEL, LOW_CONFIDENCE_THRESHOLD
from services.ai_service import OPENAI_API_KEY

logger = logging.getLogger(__name__)

GANTT_AI_SYSTEM_PROMPT = """You extract Gantt chart / project schedule data from documents.

Return ONLY a single JSON object (no markdown fences, no commentary) with this shape:
{
  "tasks": [
    {
      "temp_id": "t1",
      "type": "task" | "milestone",
      "phase": "optional phase name",
      "title": "Task title",
      "description": "optional",
      "start_date": "YYYY-MM-DD or null",
      "end_date": "YYYY-MM-DD or null",
      "status": "not_started",
      "responsible_party": "optional",
      "dependencies": [{"temp_task_id": "t0", "type": "finish_to_start"}],
      "field_confidence": {"title": 0.0-1.0, "start_date": 0.0-1.0, "end_date": 0.0-1.0},
      "warnings": ["optional per-task warnings"]
    }
  ],
  "warnings": ["global warnings"]
}

Rules:
- Use unique temp_id values (t1, t2, ...).
- Dates must be YYYY-MM-DD when confident; null when uncertain.
- finish_to_start dependencies only; temp_task_id must reference another temp_id in tasks.
- Milestones: start_date required, end_date equals start_date.
- field_confidence: 0.0-1.0 per extracted field; use <= 0.5 when guessing.
- Include warnings for ambiguous dates, missing dependencies, or unclear titles."""


def _make_temp_id(index: int) -> str:
    return f"t{index + 1}"


def _coerce_cell_str(value: Any) -> str:
    """Normalize spreadsheet cell values (including Excel dates) to strings."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _normalize_date(value: Any) -> Tuple[Optional[str], float, List[str]]:
    """Try to normalize a date string or spreadsheet date; return (iso_date, confidence, warnings)."""
    warnings: List[str] = []
    if value is None:
        return None, 0.0, warnings

    if isinstance(value, datetime):
        return value.date().isoformat(), 0.98, warnings
    if isinstance(value, date):
        return value.isoformat(), 0.98, warnings

    raw = _coerce_cell_str(value)
    if not raw:
        return None, 0.0, warnings
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d"), 0.95, warnings
        except ValueError:
            continue

    warnings.append(f"Could not parse date: {raw}")
    return None, 0.3, warnings


def _apply_low_confidence_flags(task: Dict[str, Any]) -> Dict[str, Any]:
    """Ensure low-confidence fields produce warnings in the task."""
    confidence = task.get("field_confidence") or {}
    warnings = list(task.get("warnings") or [])
    for field, score in confidence.items():
        if isinstance(score, (int, float)) and score < LOW_CONFIDENCE_THRESHOLD:
            warnings.append(f"Low confidence for {field} ({score:.0%})")
    task["warnings"] = warnings
    return task


def _normalize_ai_tasks(raw_tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    for index, raw in enumerate(raw_tasks):
        task = {
            "temp_id": raw.get("temp_id") or _make_temp_id(index),
            "type": raw.get("type", "task"),
            "phase": raw.get("phase"),
            "title": raw.get("title") or f"Untitled task {index + 1}",
            "description": raw.get("description"),
            "start_date": raw.get("start_date"),
            "end_date": raw.get("end_date"),
            "duration_days": raw.get("duration_days"),
            "status": raw.get("status", "not_started"),
            "responsible_party": raw.get("responsible_party"),
            "dependencies": raw.get("dependencies") or [],
            "field_confidence": raw.get("field_confidence") or {},
            "warnings": raw.get("warnings") or [],
        }
        normalized.append(_apply_low_confidence_flags(task))
    return normalized


def _strip_markdown_fences(text: str) -> str:
    """Remove ```json ... ``` wrappers often returned by chat models."""
    stripped = text.strip()
    if not stripped.startswith("```"):
        return stripped
    stripped = re.sub(r"^```(?:json)?\s*", "", stripped, flags=re.IGNORECASE)
    stripped = re.sub(r"\s*```$", "", stripped)
    return stripped.strip()


def _load_json_payload(response_text: str) -> Optional[Any]:
    """Parse JSON from model output, tolerating fences and leading prose."""
    if not response_text or not str(response_text).strip():
        return None

    candidates = [_strip_markdown_fences(str(response_text))]
    brace_start = candidates[0].find("{")
    bracket_start = candidates[0].find("[")
    if brace_start >= 0:
        candidates.append(candidates[0][brace_start:])
    if bracket_start >= 0:
        candidates.append(candidates[0][bracket_start:])

    decoder = json.JSONDecoder()
    for candidate in candidates:
        if not candidate:
            continue
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            for opener, closer in (("{", "}"), ("[", "]")):
                start = candidate.find(opener)
                if start < 0:
                    continue
                try:
                    payload, _ = decoder.raw_decode(candidate[start:])
                    return payload
                except json.JSONDecodeError:
                    continue
    return None


def _call_gantt_extraction_ai(client: Any, messages: List[Dict[str, Any]]) -> str:
    """Call GANTT_EXTRACTION_MODEL with JSON-object response format when supported."""
    base_kwargs: Dict[str, Any] = {
        "model": GANTT_EXTRACTION_MODEL,
        "messages": messages,
        "response_format": {"type": "json_object"},
    }
    token_attempts = (
        {"max_completion_tokens": 4000},
        {"max_tokens": 4000},
        {},
    )
    last_exc: Optional[Exception] = None
    for token_kwargs in token_attempts:
        try:
            response = client.chat.completions.create(**base_kwargs, **token_kwargs)
            return response.choices[0].message.content or ""
        except TypeError as exc:
            last_exc = exc
            continue
        except Exception as exc:
            err = str(exc).lower()
            if "response_format" in err or "json_object" in err:
                response = client.chat.completions.create(
                    model=GANTT_EXTRACTION_MODEL,
                    messages=messages,
                    **token_kwargs,
                )
                return response.choices[0].message.content or ""
            last_exc = exc
            break
    if last_exc:
        raise last_exc
    return ""


def _parse_ai_json_response(response_text: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    payload = _load_json_payload(response_text)
    if payload is None:
        logger.warning(
            "Gantt AI JSON parse failed (first 500 chars): %s",
            (response_text or "")[:500],
        )
        return [], ["AI response JSON could not be parsed"]

    if isinstance(payload, list):
        tasks_raw = payload
        warnings: List[str] = []
    elif isinstance(payload, dict):
        tasks_raw = payload.get("tasks") or payload.get("items") or []
        warnings = list(payload.get("warnings") or [])
    else:
        return [], ["AI response JSON had unexpected structure"]

    if not isinstance(tasks_raw, list):
        return [], ["AI response tasks field was not a list"]

    tasks = _normalize_ai_tasks(tasks_raw)
    return tasks, warnings


async def parse_pdf_text(file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Extract draft tasks from a PDF with a text layer via PyMuPDF + GANTT_EXTRACTION_MODEL."""
    warnings: List[str] = []
    try:
        doc = fitz.open(file_path)
        text_content = "".join(page.get_text() for page in doc)
        doc.close()
    except Exception as exc:
        logger.error("PDF text extraction failed: %s", exc)
        return [], [f"PDF read failed: {exc}"]

    if not text_content.strip():
        return [], ["PDF contains no extractable text; try uploading a scanned image instead"]

    if not OPENAI_API_KEY:
        warnings.append("AI extraction unavailable (OPENAI_API_KEY not set)")
        return [], warnings

    try:
        import openai

        client = openai.OpenAI(api_key=OPENAI_API_KEY)
        content = _call_gantt_extraction_ai(
            client,
            [
                {"role": "system", "content": GANTT_AI_SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": f"Extract Gantt tasks from this PDF text:\n\n{text_content[:8000]}",
                },
            ],
        )
        return _parse_ai_json_response(content)
    except Exception as exc:
        logger.error("PDF AI extraction failed: %s", exc)
        return [], [f"AI extraction failed: {exc}"]


async def parse_image_vision(file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Extract draft tasks from scanned Gantt / images via GANTT_EXTRACTION_MODEL vision."""
    if not OPENAI_API_KEY:
        return [], ["AI extraction unavailable (OPENAI_API_KEY not set)"]

    file_ext = os.path.splitext(file_path)[1].lower()
    mime_types = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
    }
    mime_type = mime_types.get(file_ext)
    if not mime_type:
        return [], [f"Unsupported image type: {file_ext}"]

    try:
        import base64
        import openai

        with open(file_path, "rb") as img_file:
            image_data = base64.b64encode(img_file.read()).decode("utf-8")

        client = openai.OpenAI(api_key=OPENAI_API_KEY)
        content = _call_gantt_extraction_ai(
            client,
            [
                {"role": "system", "content": GANTT_AI_SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": "Extract all Gantt chart tasks, phases, dates, and dependencies from this image.",
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:{mime_type};base64,{image_data}"},
                        },
                    ],
                },
            ],
        )
        return _parse_ai_json_response(content)
    except Exception as exc:
        logger.error("Image vision extraction failed: %s", exc)
        return [], [f"AI extraction failed: {exc}"]


def _find_column(row_keys: List[str], candidates: List[str]) -> Optional[str]:
    normalized = {k.lower().strip(): k for k in row_keys}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
    return None


async def parse_csv_milestones(file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Deterministic CSV milestone/task row mapping."""
    global_warnings: List[str] = []
    tasks: List[Dict[str, Any]] = []

    try:
        with open(file_path, newline="", encoding="utf-8-sig") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(handle, dialect=dialect)
            if not reader.fieldnames:
                return [], ["CSV has no header row"]

            fieldnames = list(reader.fieldnames)
            for index, row in enumerate(reader):
                task = _draft_task_from_row_dict(row, index, fieldnames)
                if task:
                    tasks.append(task)

    except Exception as exc:
        logger.error("CSV parse failed: %s", exc)
        return [], [f"CSV parse failed: {exc}"]

    if not tasks:
        global_warnings.append("No tasks found in CSV")
        return tasks, global_warnings

    return _resolve_draft_dependencies(tasks), global_warnings


def _draft_task_from_row_dict(
    row: Dict[str, Any],
    index: int,
    fieldnames: List[str],
) -> Dict[str, Any]:
    """Map a header→value row dict to a GanttDraftTask-shaped dict."""
    title_col = _find_column(
        fieldnames,
        ["title", "name", "task", "milestone", "task name", "activity"],
    )
    if not title_col or not (row.get(title_col) or "").strip():
        return {}

    phase_col = _find_column(fieldnames, ["phase", "stage", "group"])
    start_col = _find_column(fieldnames, ["start_date", "start", "begin", "start date"])
    end_col = _find_column(fieldnames, ["end_date", "end", "finish", "due", "end date"])
    type_col = _find_column(fieldnames, ["type"])
    desc_col = _find_column(fieldnames, ["description", "notes", "details"])
    status_col = _find_column(fieldnames, ["status"])
    resp_col = _find_column(
        fieldnames,
        ["responsible_party", "owner", "assignee", "responsible"],
    )
    dep_col = _find_column(
        fieldnames,
        ["dependencies", "depends_on", "predecessor", "predecessors"],
    )
    task_id_col = _find_column(fieldnames, ["task_id", "id"])

    raw_type = _coerce_cell_str(row.get(type_col) if type_col else "task").lower() or "task"
    task_type = "milestone" if raw_type in {"milestone", "m"} else "task"

    start_date, start_conf, start_warn = _normalize_date(
        row.get(start_col) if start_col else None
    )
    end_date, end_conf, end_warn = _normalize_date(
        row.get(end_col) if end_col else None
    )

    field_confidence = {"title": 0.98}
    task_warnings = start_warn + end_warn

    if start_col and start_date is None and row.get(start_col):
        field_confidence["start_date"] = start_conf
    elif start_date:
        field_confidence["start_date"] = start_conf

    if end_col and end_date is None and row.get(end_col):
        field_confidence["end_date"] = end_conf
    elif end_date:
        field_confidence["end_date"] = end_conf

    if task_type == "milestone" and start_date:
        end_date = start_date

    dependencies: List[Dict[str, str]] = []
    if dep_col and row.get(dep_col):
        dep_refs = [p.strip() for p in re.split(r"[;,|]", str(row[dep_col])) if p.strip()]
        for ref in dep_refs:
            dependencies.append({"temp_task_id": ref, "type": "finish_to_start"})
        field_confidence["dependencies"] = 0.7
        task_warnings.append("Dependency references need review after import")

    exported_task_id = _coerce_cell_str(row.get(task_id_col)) if task_id_col else ""

    task = {
        "temp_id": _make_temp_id(index),
        "type": task_type,
        "phase": _coerce_cell_str(row.get(phase_col)) or None if phase_col else None,
        "title": _coerce_cell_str(row[title_col]),
        "description": _coerce_cell_str(row.get(desc_col)) or None if desc_col else None,
        "start_date": start_date,
        "end_date": end_date,
        "status": _coerce_cell_str(row.get(status_col)) or "not_started" if status_col else "not_started",
        "responsible_party": _coerce_cell_str(row.get(resp_col)) or None if resp_col else None,
        "dependencies": dependencies,
        "field_confidence": field_confidence,
        "warnings": task_warnings,
    }
    if exported_task_id:
        task["task_id"] = exported_task_id
    return _apply_low_confidence_flags(task)


def _resolve_draft_dependencies(tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Resolve dependency references (title, temp_id, or exported task_id) to temp_ids."""
    title_to_temp: Dict[str, str] = {}
    task_id_to_temp: Dict[str, str] = {}
    for task in tasks:
        title_key = task["title"].strip().lower()
        if title_key and title_key not in title_to_temp:
            title_to_temp[title_key] = task["temp_id"]
        exported_id = (task.get("task_id") or "").strip()
        if exported_id:
            task_id_to_temp[exported_id] = task["temp_id"]

    for task in tasks:
        resolved: List[Dict[str, str]] = []
        for dep in task.get("dependencies") or []:
            ref = dep.get("temp_task_id", "").strip()
            if not ref:
                continue
            ref_lower = ref.lower()
            target = None
            if ref in {t["temp_id"] for t in tasks}:
                target = ref
            elif ref in task_id_to_temp:
                target = task_id_to_temp[ref]
            elif ref_lower in title_to_temp:
                target = title_to_temp[ref_lower]
            if target and target != task["temp_id"]:
                resolved.append({"temp_task_id": target, "type": "finish_to_start"})
            elif not target:
                task.setdefault("warnings", []).append(
                    f"Unresolved dependency reference: {ref}"
                )
                task.setdefault("field_confidence", {})["dependencies"] = 0.3
        task["dependencies"] = resolved
        task["warnings"] = _apply_low_confidence_flags(task)["warnings"]
    return tasks


async def parse_excel(file_path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Deterministic Excel (.xlsx) task row mapping (Tasks sheet or first sheet)."""
    global_warnings: List[str] = []
    tasks: List[Dict[str, Any]] = []

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb["Tasks"] if "Tasks" in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], ["Excel sheet has no header row"]

        fieldnames = [
            str(cell).strip() if cell is not None else ""
            for cell in header_row
        ]
        if not any(fieldnames):
            return [], ["Excel sheet has no header row"]

        for index, row_values in enumerate(rows_iter):
            if not row_values or not any(v is not None and str(v).strip() for v in row_values):
                continue
            row_dict = {
                fieldnames[i]: row_values[i] if i < len(row_values) else None
                for i in range(len(fieldnames))
            }
            task = _draft_task_from_row_dict(row_dict, index, fieldnames)
            if task:
                tasks.append(task)

        wb.close()
    except Exception as exc:
        logger.error("Excel parse failed: %s", exc)
        return [], [f"Excel parse failed: {exc}"]

    if not tasks:
        global_warnings.append("No tasks found in Excel file")
        return tasks, global_warnings

    return _resolve_draft_dependencies(tasks), global_warnings


def make_temp_id() -> str:
    """Generate a unique temp_id for manually added draft rows."""
    return f"t_{uuid.uuid4().hex[:8]}"
