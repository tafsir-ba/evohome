"""
Gantt export/import unit tests (no DB, no API).
"""
import asyncio
import importlib.util
import os
import sys
import tempfile
import types
from pathlib import Path

import pytest

_services_dir = Path(__file__).resolve().parents[1] / "services"

if "services" not in sys.modules:
    _services_pkg = types.ModuleType("services")
    _services_pkg.__path__ = [str(_services_dir)]
    sys.modules["services"] = _services_pkg

sys.modules.setdefault("fitz", types.ModuleType("fitz"))
_ai_service = types.ModuleType("services.ai_service")
_ai_service.OPENAI_API_KEY = ""
sys.modules["services.ai_service"] = _ai_service

_export_path = _services_dir / "gantt_export_service.py"
_export_spec = importlib.util.spec_from_file_location(
    "services.gantt_export_service", _export_path
)
_gantt_export = importlib.util.module_from_spec(_export_spec)
sys.modules["services.gantt_export_service"] = _gantt_export
_export_spec.loader.exec_module(_gantt_export)

_parsers_path = _services_dir / "gantt_parsers.py"
_parsers_spec = importlib.util.spec_from_file_location(
    "services.gantt_parsers", _parsers_path
)
_gantt_parsers = importlib.util.module_from_spec(_parsers_spec)
sys.modules["services.gantt_parsers"] = _gantt_parsers
_parsers_spec.loader.exec_module(_gantt_parsers)

build_csv_content = _gantt_export.build_csv_content
build_xlsx_bytes = _gantt_export.build_xlsx_bytes
build_pdf_bytes = _gantt_export.build_pdf_bytes
parse_excel = _gantt_parsers.parse_excel

SAMPLE_PROJECT = {"title": "Test Project / Q2", "description": "Demo"}
SAMPLE_TASKS = [
    {
        "order": 0,
        "task_id": "gt_pre",
        "type": "task",
        "phase": "Phase 1",
        "title": "Predecessor",
        "description": None,
        "start_date": "2026-02-20",
        "end_date": "2026-02-28",
        "duration_days": 8,
        "status": "completed",
        "responsible_party": None,
        "dependencies": [],
        "source": "manual",
    },
    {
        "order": 1,
        "task_id": "gt_succ",
        "type": "task",
        "phase": "Phase 1",
        "title": "Foundation",
        "description": "Pour concrete",
        "start_date": "2026-03-01",
        "end_date": "2026-03-05",
        "duration_days": 4,
        "status": "not_started",
        "responsible_party": "Team A",
        "dependencies": [{"task_id": "gt_pre", "type": "finish_to_start"}],
        "source": "manual",
    },
    {
        "order": 2,
        "type": "milestone",
        "phase": "Phase 1",
        "title": "Kickoff",
        "description": None,
        "start_date": "2026-02-01",
        "end_date": "2026-02-01",
        "duration_days": 0,
        "status": "completed",
        "responsible_party": None,
        "dependencies": [],
        "source": "manual",
    },
]


class TestGanttExport:
    def test_csv_columns_match_spec(self):
        csv_text = build_csv_content(SAMPLE_TASKS)
        header = csv_text.splitlines()[0]
        assert header == (
            "order,type,phase,title,description,start_date,end_date,"
            "duration_days,status,responsible_party,dependencies,source"
        )
        assert "gt_pre" in csv_text

    def test_xlsx_has_tasks_and_metadata_sheets(self):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        data = build_xlsx_bytes(SAMPLE_PROJECT, SAMPLE_TASKS)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(data)
            path = tmp.name
        try:
            wb = load_workbook(path, read_only=True)
            assert "Tasks" in wb.sheetnames
            assert "Metadata" in wb.sheetnames
            tasks_ws = wb["Tasks"]
            rows = list(tasks_ws.iter_rows(values_only=True))
            assert rows[0][3] == "title"
            assert any(row[3] == "Foundation" for row in rows[1:])
            meta_ws = wb["Metadata"]
            assert meta_ws["B1"].value == "Test Project / Q2"
            wb.close()
        finally:
            os.unlink(path)

    def test_pdf_bytes_non_empty(self):
        pytest.importorskip("reportlab")
        data = build_pdf_bytes(SAMPLE_PROJECT, SAMPLE_TASKS)
        assert data[:4] == b"%PDF"

    def test_presentation_pdf_fits_one_page_for_small_project(self):
        pytest.importorskip("reportlab")
        import re

        data = build_pdf_bytes(SAMPLE_PROJECT, SAMPLE_TASKS, mode="presentation")
        assert data[:4] == b"%PDF"
        assert len(re.findall(rb"/Type /Page\n", data)) == 1
        assert b"/MediaBox [ 0 0 1190.551 841.8898 ]" in data  # A3 landscape

    def test_detailed_pdf_mode(self):
        pytest.importorskip("reportlab")
        from services.gantt_export_service import build_pdf_response

        _, _, disposition = build_pdf_response(SAMPLE_PROJECT, SAMPLE_TASKS, mode="detailed")
        assert "_detailed.pdf" in disposition

    def test_excel_import_round_trip(self):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        data = build_xlsx_bytes(SAMPLE_PROJECT, SAMPLE_TASKS)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(data)
            path = tmp.name
        try:
            tasks, warnings = asyncio.run(parse_excel(path))
            assert not warnings or "No tasks" not in " ".join(warnings)
            titles = {t["title"] for t in tasks}
            assert "Foundation" in titles
            assert "Kickoff" in titles
        finally:
            os.unlink(path)

    def test_excel_import_native_date_cells(self):
        pytest.importorskip("openpyxl")
        from datetime import date
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Tasks"
            ws.append(["title", "start_date", "end_date", "type"])
            ws.append(["Dated Task", date(2026, 5, 10), date(2026, 5, 15), "task"])
            wb.save(path)
            wb.close()

            tasks, warnings = asyncio.run(parse_excel(path))
            assert len(tasks) == 1
            assert tasks[0]["start_date"] == "2026-05-10"
            assert tasks[0]["end_date"] == "2026-05-15"
            assert not any("Could not parse date" in w for w in tasks[0].get("warnings", []))
        finally:
            os.unlink(path)

    def test_excel_import_resolves_exported_task_id_dependencies(self):
        pytest.importorskip("openpyxl")

        data = build_xlsx_bytes(SAMPLE_PROJECT, SAMPLE_TASKS)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(data)
            path = tmp.name
        try:
            tasks, _ = asyncio.run(parse_excel(path))
            foundation = next(t for t in tasks if t["title"] == "Foundation")
            assert foundation["dependencies"]
            assert foundation["dependencies"][0]["temp_task_id"] != "gt_pre"
        finally:
            os.unlink(path)

    def test_excel_import_from_tasks_sheet(self):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Tasks"
            ws.append(
                [
                    "order",
                    "type",
                    "phase",
                    "title",
                    "start_date",
                    "end_date",
                    "status",
                ]
            )
            ws.append([0, "task", "P1", "Excel Task", "2026-04-01", "2026-04-03", "not_started"])
            wb.save(path)
            wb.close()

            tasks, _ = asyncio.run(parse_excel(path))
            assert len(tasks) == 1
            assert tasks[0]["title"] == "Excel Task"
            assert tasks[0]["start_date"] == "2026-04-01"
        finally:
            os.unlink(path)
